import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from pathlib import Path
from .violation_data_model import ViolationReportData

class ViolationReportGenerator:
    """
    Генерирует отчет о нарушениях в формате Excel.
    Стиль похож на основной отчет, но без оценок и с колонкой для фото.
    """
    
    def generate(self, data: ViolationReportData, output_path: Path) -> Path:
        """Генерирует отчет о нарушениях"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Отчет о нарушениях"
        
        # Настройка ширины столбцов
        self._setup_column_widths(ws)
        
        # Заполнение отчета
        row_idx = self._setup_header(ws, data)
        row_idx = self._setup_table_header(ws, row_idx)
        row_idx = self._fill_violations_table(ws, data.violations, row_idx)
        self._add_summary(ws, data.total_violations, row_idx)
        
        # Применение стилей
        self._apply_styles(ws, row_idx)
        
        wb.save(output_path)
        return output_path
    
    def _setup_column_widths(self, ws):
        """Настраивает ширину столбцов"""
        column_widths = {
            'A': 8,    # № п/п
            'B': 25,   # Раздел
            'C': 25,   # Подраздел
            'D': 15,   # № нарушения
            'E': 50,   # Комментарий
            'F': 30,   # Фото нарушения
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
    
    def _setup_header(self, ws, data: ViolationReportData) -> int:
        """Создает заголовок отчета"""
        # Основной заголовок
        ws.merge_cells('A1:F1')
        ws['A1'] = "ОТЧЕТ О ВЫЯВЛЕННЫХ НАРУШЕНИЯХ"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Информация о проверке
        ws['A2'] = f"Структурное подразделение: {data.section_name}"
        ws['A2'].font = Font(bold=True, size=12)
        
        if data.inspection_date:
            ws['A3'] = f"Дата проведения проверки: {data.inspection_date}"
        else:
            ws['A3'] = "Дата проведения проверки: не указана"
        ws['A3'].font = Font(bold=True, size=12)
        
        if data.inspector:
            ws['A4'] = f"Проверяющий: {data.inspector}"
        else:
            ws['A4'] = "Проверяющий: не указан"
        ws['A4'].font = Font(bold=True, size=12)
        
        ws['A5'] = f"Исходный отчет: {data.original_report_name}"
        ws['A5'].font = Font(italic=True)
        
        return 7  # Возвращаем следующую строку для заполнения
    
    def _setup_table_header(self, ws, start_row: int) -> int:
        """Создает заголовок таблицы"""
        # Заголовок таблицы
        ws.merge_cells(f'A{start_row}:F{start_row}')
        ws.cell(row=start_row, column=1).value = "Перечень выявленных нарушений"
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=14)
        ws.cell(row=start_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        
        start_row += 1
        
        # Заголовки столбцов
        headers = [
            '№ п/п', 'Раздел', 'Подраздел', '№ нарушения', 
            'Комментарий', 'Фотография нарушения'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
        
        return start_row + 1
    
    def _fill_violations_table(self, ws, violations, start_row: int) -> int:
        """Заполняет таблицу нарушениями"""
        current_row = start_row
        
        for i, violation in enumerate(violations, 1):
            # № п/п
            ws.cell(row=current_row, column=1).value = i
            # Раздел
            ws.cell(row=current_row, column=2).value = violation.section_name
            # Подраздел
            ws.cell(row=current_row, column=3).value = violation.subsection_name or "-"
            # № нарушения
            ws.cell(row=current_row, column=4).value = violation.criterion_number
            # Комментарий
            ws.cell(row=current_row, column=5).value = violation.comment
            
            # Обработка фотографии
            if violation.photo_path and violation.photo_path.exists():
                self._insert_photo(ws, violation.photo_path, current_row, 6)
            else:
                ws.cell(row=current_row, column=6).value = "[Фото отсутствует]"
                ws.cell(row=current_row, column=6).font = Font(italic=True, color="808080")
            
            # Устанавливаем высоту строки для размещения фото
            ws.row_dimensions[current_row].height = 80
            
            current_row += 1
        
        return current_row
    
    def _insert_photo(self, ws, photo_path: Path, row: int, col: int):
        """Вставляет фотографию в указанную ячейку"""
        try:
            img = Image(photo_path)
            
            # Масштабируем изображение
            max_width = 200
            max_height = 70
            
            # Получаем оригинальные размеры
            original_width = img.width
            original_height = img.height
            
            # Вычисляем коэффициенты масштабирования
            width_ratio = max_width / original_width
            height_ratio = max_height / original_height
            scale_ratio = min(width_ratio, height_ratio)
            
            # Применяем масштабирование
            img.width = int(original_width * scale_ratio)
            img.height = int(original_height * scale_ratio)
            
            # Якорь для позиционирования изображения
            cell_anchor = f'{get_column_letter(col)}{row}'
            
            # Добавляем изображение
            img.anchor = cell_anchor
            ws.add_image(img)
            
        except Exception as e:
            # Если не удалось вставить фото, пишем сообщение
            ws.cell(row=row, column=col).value = f"[Ошибка загрузки фото: {e}]"
            ws.cell(row=row, column=col).font = Font(italic=True, color="FF0000")
    
    def _add_summary(self, ws, total_violations: int, row_idx: int):
        """Добавляет итоговую информацию"""
        summary_row = row_idx + 2
        
        ws.merge_cells(f'A{summary_row}:F{summary_row}')
        ws.cell(row=summary_row, column=1).value = (
            f"Всего выявлено нарушений: {total_violations}"
        )
        ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=summary_row, column=1).alignment = Alignment(horizontal='center')
        
        # Добавляем блок для подписи (как в основном отчете)
        signature_row = summary_row + 2
        ws.merge_cells(f'A{signature_row}:F{signature_row}')
        ws.cell(row=signature_row, column=1).value = "Проверку проводил _____________________ ______________ _______________________"
        
        signature_row += 1
        ws.merge_cells(f'A{signature_row}:F{signature_row}')
        ws.cell(row=signature_row, column=1).value = "должность подпись расшифровка подпись"
        
        # Стиль для блока подписи
        for row in range(signature_row-1, signature_row+1):
            ws.row_dimensions[row].height = 25
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center')
    
    def _apply_styles(self, ws, last_data_row: int):
        """Применяет стили к отчету"""
        # Стили
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        light_red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
        # Заголовок таблицы (последняя строка заголовка)
        header_row = 7  # Строка с заголовками столбцов
        for col in range(1, 7):
            cell = ws.cell(row=header_row, column=col)
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = border
            cell.fill = header_fill
        
        # Данные нарушений
        for row in range(header_row + 1, last_data_row):
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                # Заливка для четных строк
                if row % 2 == 0:
                    cell.fill = light_red_fill
                
                # Выравнивание
                if col in [1, 4]:  # № п/п и № нарушения
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align