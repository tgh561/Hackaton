import json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime
from typing import Dict, Optional

class MonthlyReportGenerator:
    """
    Генератор сводного отчета за месяц по культуре производства.
    Обновляет данные ежедневно и рассчитывает динамику.
    """
    
    def __init__(self, template_path: Path, output_dir: Path):
        self.template_path = template_path
        self.output_dir = output_dir
        self.monthly_data = self._load_template()
        self.current_month = datetime.now().strftime("%Y-%m")
        
    def _load_template(self) -> Dict:
        """Загружает шаблон отчета из JSON"""
        with open(self.template_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    
    def add_daily_report(self, department_id: str, day: int, score: float) -> bool:
        """
        Добавляет ежедневную оценку для подразделения
        
        Args:
            department_id: ID подразделения (например, "1.1")
            day: Число месяца (1-31)
            score: Оценка за день
            
        Returns:
            bool: Успешно ли добавлена оценка
        """
        for dept in self.monthly_data["departments"]:
            if dept["id"] == department_id and dept["type"] == "leaf":
                if str(day) in dept["daily_scores"]:
                    dept["daily_scores"][str(day)] = score
                    self._recalculate_monthly_scores()
                    return True
        return False
    
    def _recalculate_monthly_scores(self):
        """Пересчитывает средние оценки за месяц для всех подразделений"""
        # Сначала пересчитываем листовые подразделения
        for dept in self.monthly_data["departments"]:
            if dept["type"] == "leaf":
                scores = [s for s in dept["daily_scores"].values() if s is not None]
                if scores:
                    dept["monthly_score"]["current"] = round(sum(scores) / len(scores), 1)
                    dept["monthly_score"]["dynamics"] = (
                        dept["monthly_score"]["current"] - dept["monthly_score"]["previous"]
                        if dept["monthly_score"]["current"] is not None 
                        and dept["monthly_score"]["previous"] is not None 
                        else None
                    )
        
        # Затем пересчитываем агрегированные подразделения
        for dept in self.monthly_data["departments"]:
            if dept["type"] == "aggregated":
                child_scores = []
                for child_id in dept["aggregated_from"]:
                    child = self._get_department_by_id(child_id)
                    if child and child["monthly_score"]["current"] is not None:
                        child_scores.append(child["monthly_score"]["current"])
                
                if child_scores:
                    dept["monthly_score"]["current"] = round(sum(child_scores) / len(child_scores), 1)
                    dept["monthly_score"]["dynamics"] = (
                        dept["monthly_score"]["current"] - dept["monthly_score"]["previous"]
                        if dept["monthly_score"]["current"] is not None 
                        and dept["monthly_score"]["previous"] is not None 
                        else None
                    )
    
    def _get_department_by_id(self, department_id: str) -> Optional[Dict]:
        """Находит подразделение по ID"""
        for dept in self.monthly_data["departments"]:
            if dept["id"] == department_id:
                return dept
        return None
    
    def generate_excel_report(self) -> Path:
        """Генерирует Excel отчет"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Сводная аналитика"
        
        self._setup_column_widths(ws)
        self._setup_header(ws)
        self._fill_data(ws)
        self._apply_styles(ws)
        
        output_path = self.output_dir / f"сводный_отчет_{self.current_month}.xlsx"
        wb.save(output_path)
        return output_path
    
    def _setup_column_widths(self, ws):
        """Настраивает ширину столбцов"""
        column_widths = {
            'A': 8,    # № п/п
            'B': 25,   # Структурное подразделение
            'C': 15,   # Форма чек-листа
            'D': 40,   # Пояснения
            'E': 12,   # Оценка за день
        }
        
        # Добавляем колонки для каждого дня
        for i, day in enumerate(self.monthly_data["days"], 6):
            col_letter = get_column_letter(i)
            column_widths[col_letter] = 8
        
        # Колонки для итогов
        column_widths[get_column_letter(len(self.monthly_data["days"]) + 6)] = 15  # Текущий месяц
        column_widths[get_column_letter(len(self.monthly_data["days"]) + 7)] = 15  # Предыдущий месяц
        column_widths[get_column_letter(len(self.monthly_data["days"]) + 8)] = 12  # Динамика
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
    
    def _setup_header(self, ws):
        """Создает заголовок отчета"""
        # Основной заголовок
        last_col = get_column_letter(len(self.monthly_data["days"]) + 8)
        ws.merge_cells(f'A1:{last_col}1')
        ws['A1'] = self.monthly_data["title"]
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Заголовки столбцов
        headers = [
            '№ п/п', 'Структурное подразделение', 'ФОРМА ЧЕК-ЛИСТА', 'ПОЯСНЕНИЯ', 'ОЦЕНКА ЗА ДЕНЬ'
        ]
        
        # Добавляем заголовки дней
        headers.extend([str(day) for day in self.monthly_data["days"]])
        
        # Добавляем итоговые заголовки
        headers.extend([
            'ИТОГОВАЯ ОЦЕНКА ЗА МЕСЯЦ (ДИНАМИКА)',
            '',
            ''
        ])
        
        # Записываем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
        
        # Подзаголовки для итоговой оценки
        ws.cell(row=4, column=len(headers)-2).value = "за текущий месяц"
        ws.cell(row=4, column=len(headers)-1).value = "за предыдущий месяц"
        ws.cell(row=4, column=len(headers)).value = "динамика"
        
        # Подзаголовок "Число месяца"
        ws.merge_cells(f'E4:{get_column_letter(len(self.monthly_data["days"]) + 4)}4')
        ws.cell(row=4, column=5).value = "Число месяца"
        ws.cell(row=4, column=5).alignment = Alignment(horizontal='center')
    
    def _fill_data(self, ws):
        """Заполняет таблицу данными"""
        current_row = 5
        
        for dept in self.monthly_data["departments"]:
            # № п/п
            ws.cell(row=current_row, column=1).value = dept["id"]
            # Структурное подразделение
            ws.cell(row=current_row, column=2).value = dept["name"]
            # Форма чек-листа
            ws.cell(row=current_row, column=3).value = dept["form_type"] or ""
            # Пояснения
            ws.cell(row=current_row, column=4).value = dept["description"]
            
            # Оценки за дни
            col = 5
            for day in self.monthly_data["days"]:
                score = dept.get("daily_scores", {}).get(str(day))
                if dept["type"] == "aggregated" and score is None:
                    # Для агрегированных подразделений вычисляем среднее за день
                    child_scores = []
                    for child_id in dept["aggregated_from"]:
                        child = self._get_department_by_id(child_id)
                        if child and child.get("daily_scores", {}).get(str(day)) is not None:
                            child_scores.append(child["daily_scores"][str(day)])
                    
                    if child_scores:
                        score = round(sum(child_scores) / len(child_scores), 1)
                
                ws.cell(row=current_row, column=col).value = score
                col += 1
            
            # Итоговые оценки
            last_data_col = len(self.monthly_data["days"]) + 4
            ws.cell(row=current_row, column=last_data_col + 1).value = dept["monthly_score"]["current"]
            ws.cell(row=current_row, column=last_data_col + 2).value = dept["monthly_score"]["previous"]
            ws.cell(row=current_row, column=last_data_col + 3).value = dept["monthly_score"]["dynamics"]
            
            current_row += 1
        
        # Добавляем примечания
        current_row += 2
        for note in self.monthly_data["notes"]:
            ws.merge_cells(f'A{current_row}:{get_column_letter(len(self.monthly_data["days"]) + 8)}{current_row}')
            ws.cell(row=current_row, column=1).value = note
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left', wrap_text=True)
            current_row += 1
    
    def _apply_styles(self, ws):
        """Применяет стили к отчету"""
        # Стили
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Стили для заголовков
        for row in [1, 3, 4]:
            for col in range(1, len(self.monthly_data["days"]) + 9):
                cell = ws.cell(row=row, column=col)
                if row in [3, 4]:
                    cell.font = bold_font
                    cell.fill = header_fill
                cell.border = border
                cell.alignment = center_align
        
        # Стили для данных
        last_row = 5 + len(self.monthly_data["departments"])
        for row in range(5, last_row):
            for col in range(1, len(self.monthly_data["days"]) + 9):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col >= 5 and col <= len(self.monthly_data["days"]) + 4:
                    # Оценки за дни
                    cell.alignment = center_align
                elif col > len(self.monthly_data["days"]) + 4:
                    # Итоговые колонки
                    cell.alignment = center_align
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        if col == len(self.monthly_data["days"]) + 8:  # Динамика
                            if cell.value > 0:
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            elif cell.value < 0:
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    def save_monthly_data(self):
        """Сохраняет обновленные данные обратно в JSON"""
        with open(self.template_path, 'w', encoding='utf-8') as f:
            json.dump(self.monthly_data, f, ensure_ascii=False, indent=2)
    
    def get_department_stats(self, department_id: str) -> Dict:
        """Возвращает статистику по подразделению"""
        dept = self._get_department_by_id(department_id)
        if dept:
            filled_days = sum(1 for score in dept.get("daily_scores", {}).values() if score is not None)
            total_days = len(dept.get("daily_scores", {}))
            return {
                "department": dept["name"],
                "filled_days": filled_days,
                "total_days": total_days,
                "completion_percentage": round((filled_days / total_days) * 100, 1) if total_days > 0 else 0,
                "current_score": dept["monthly_score"]["current"],
                "previous_score": dept["monthly_score"]["previous"],
                "dynamics": dept["monthly_score"]["dynamics"]
            }
        return {}


class DailyReportProcessor:
    """
    Обработчик ежедневных отчетов для автоматического обновления сводного отчета
    """
    
    def __init__(self, monthly_generator: MonthlyReportGenerator, daily_reports_dir: Path):
        self.monthly_generator = monthly_generator
        self.daily_reports_dir = daily_reports_dir
    
    def process_daily_reports(self, target_day: Optional[int] = None):
        """
        Обрабатывает все ежедневные отчеты за указанный день
        
        Args:
            target_day: День месяца (если None, используется текущий день)
        """
        if target_day is None:
            target_day = datetime.now().day
        
        daily_reports = list(self.daily_reports_dir.glob(f"*_report.json"))
        
        for report_path in daily_reports:
            try:
                with open(report_path, 'r', encoding='utf-8') as f:
                    report_data = json.load(f)
                
                # Извлекаем оценку из отчета
                overall_score = report_data.get("overall_score")
                section_name = report_data.get("section_name")
                
                if overall_score is not None and section_name:
                    # Сопоставляем имя раздела с ID подразделения
                    department_id = self._map_section_to_department(section_name)
                    if department_id:
                        self.monthly_generator.add_daily_report(department_id, target_day, overall_score)
                        print(f"Добавлена оценка {overall_score} для {section_name} ({department_id}) за день {target_day}")
                
            except Exception as e:
                print(f"Ошибка обработки отчета {report_path}: {e}")
        
        # Сохраняем обновленные данные
        self.monthly_generator.save_monthly_data()
    
    def _map_section_to_department(self, section_name: str) -> Optional[str]:
        """Сопоставляет имя раздела с ID подразделения"""
        mapping = {
            "УПП": "1",
            "Отделение раздува": "1.1",
            "Отделение литья": "1.2", 
            "Дробильное отделение": "1.3",
            "Помещение централизованной подачи материала": "1.4",
            "Сборочный участок": "2",
            "Участок производства СКПГ": "3",
            "Инструментальный участок": "5",
            "ЭМО": "6",
            "Помещения ЭМО": "6.1",
            "Состояние оборудования на участках": "6.2"
        }
        return mapping.get(section_name)
