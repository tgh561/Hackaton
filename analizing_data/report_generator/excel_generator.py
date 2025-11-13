import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Dict
from data_model import ChecklistData

class ExcelGenerator:
    def generate(self, data: ChecklistData, output_path: Path) -> Path:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Отчет проверки"
        
        # Настройка ширины столбцов
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 40
        
        self._setup_header(ws, data)
        row_idx = self._setup_table_header(ws)

        for section_key, section in data.sections.items():
            row_idx = self._add_section(ws, section_key, section, row_idx)

        self._add_overall_score(ws, data.overall_score, row_idx)
        row_idx += 3
        self._add_signature_block(ws, row_idx)
        self._add_footer_note(ws, row_idx + 4)
        
        self._apply_styles(ws)
        wb.save(output_path)
        return output_path

    def _setup_header(self, ws, data: ChecklistData):
        # Стиль для заголовков
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=12)
        
        if data.revision_date:
            ws['A1'] = f"Редакция от {data.revision_date}"
            ws['A1'].font = Font(italic=True)
        
        ws['A2'] = data.section_name
        ws['A2'].font = title_font
        
        ws['C2'] = "Дата проведения проверки:"
        ws['C2'].font = header_font
        if data.inspection_date:
            ws['D2'] = data.inspection_date
            
        if data.inspector:
            ws['A3'] = f"Проверяющий: {data.inspector}"
            ws['A3'].font = header_font

    def _setup_table_header(self, ws):
        # Основной заголовок таблицы
        ws.merge_cells('A5:E5')
        ws['A5'] = "Чек-лист оценки состояния оборудования и рабочего пространства рабочего персонала"
        ws['A5'].font = Font(bold=True, size=14)
        ws['A5'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Заголовки столбцов
        headers = ['№ п/п', 'Критерий оценки', 'Оценка', '', 'Краткий комментарий с указанием номера единицы оборудования, где выявлено несоответствие, и фото несоответствия']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col)
            cell.value = header
        
        # Подзаголовки для оценки
        ws['C7'] = "соответствует"
        ws['D7'] = "не соответствует"
        
        return 8

    def _add_section(self, ws, key: str, section, row_idx: int) -> int:
        # Заголовок раздела
        ws.cell(row=row_idx, column=1).value = key
        ws.cell(row=row_idx, column=2).value = section.description
        row_idx += 1

        if section.subdivisions:
            for sub_key, sub in section.subdivisions.items():
                row_idx = self._add_subdivision(ws, sub_key, sub, row_idx)
        else:
            row_idx = self._add_criteria(ws, section.criteria, row_idx)

        if section.total_score is not None:
            ws.cell(row=row_idx, column=2).value = f"Общий балл за раздел {key}"
            ws.cell(row=row_idx, column=3).value = round(section.total_score, 2)
            row_idx += 2
        
        return row_idx

    def _add_subdivision(self, ws, key: str, sub, row_idx: int) -> int:
        # Заголовок подраздела
        ws.cell(row=row_idx, column=1).value = key
        ws.cell(row=row_idx, column=2).value = sub.description
        row_idx += 1
        row_idx = self._add_criteria(ws, sub.criteria, row_idx)
        
        if sub.total_score is not None:
            ws.cell(row=row_idx, column=2).value = f"Общий балл за подраздел {key}"
            ws.cell(row=row_idx, column=3).value = round(sub.total_score, 2)
            row_idx += 2
        
        return row_idx

    def _add_criteria(self, ws, criteria, row_idx: int) -> int:
        for crit in criteria:
            ws.cell(row=row_idx, column=1).value = crit.number
            ws.cell(row=row_idx, column=2).value = crit.description
            ws.cell(row=row_idx, column=3).value = "✓" if crit.complies == 1 else ""
            ws.cell(row=row_idx, column=4).value = "✓" if crit.does_not_comply == 0 else ""
            ws.cell(row=row_idx, column=5).value = crit.comment
            row_idx += 1
        return row_idx

    def _add_overall_score(self, ws, score: float, row_idx: int):
        ws.merge_cells(f'B{row_idx}:D{row_idx}')
        ws.cell(row=row_idx, column=2).value = "Итоговая оценка структурному подразделению"
        ws.cell(row=row_idx, column=5).value = round(score, 2) if score is not None else 0

    def _add_signature_block(self, ws, row_idx: int):
        """Добавляет блок для подписи"""
        ws.merge_cells(f'A{row_idx}:E{row_idx}')
        ws.cell(row=row_idx, column=1).value = "Проверку проводил _____________________ ______________ _______________________"
        
        row_idx += 1
        ws.merge_cells(f'A{row_idx}:E{row_idx}')
        ws.cell(row=row_idx, column=1).value = "должность подпись расшифровка подпись"
        
        # Стиль для блока подписи
        for row in range(row_idx-1, row_idx+1):
            ws.row_dimensions[row].height = 25
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center')

    def _add_footer_note(self, ws, row_idx: int):
        """Добавляет примечание внизу"""
        note_text = (
            "* Итоговая оценка структурному подразделению проставляется проверяющим при выявлении одного и того же "
            "несоответствия 2 раза в размере «3 балла», при выявлении одного и того же несоответствия более 2 раз - «2 балла». "
            "При отсутствии повторяющихся несоответствий в ходе проведения проверки данная графа проверяющим не заполняется."
        )
        
        ws.merge_cells(f'A{row_idx}:E{row_idx+2}')  # Объединяем для многострочного текста
        cell = ws.cell(row=row_idx, column=1)
        cell.value = note_text
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.font = Font(italic=True, size=9)
        
        # Устанавливаем высоту строки для примечания
        ws.row_dimensions[row_idx].height = 60

    def _apply_styles(self, ws):
        # Определяем стили
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Применяем стили к заголовкам таблицы
        for row in ws['A6:E7']:
            for cell in row:
                cell.font = bold_font
                cell.alignment = center_align
                cell.border = border
                cell.fill = header_fill
        
        # Применяем границы ко всем заполненным ячейкам
        for row in ws.iter_rows(min_row=5):
            for cell in row:
                if cell.value is not None:
                    cell.border = border
                    # Выравнивание в зависимости от столбца
                    if cell.column in [1, 3, 4]:  # № п/п, соответствует, не соответствует
                        cell.alignment = center_align
                    else:
                        cell.alignment = left_align
        
        # Специальные стили для итоговой оценки
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "Итоговая оценка" in str(cell.value):
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")